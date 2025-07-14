#!/usr/bin/env python3
"""
EXCEL SKU APPROVAL SYSTEM v5 - Formula-Powered Dependent Dropdowns

Business-Optimized SKU Approval Workflow using formula-based dependent dropdowns.
This version is macro-free, enhancing compatibility and reliability.
"""

import sys
import os
import re
import json
from pathlib import Path
from typing import Dict, List, Any, Optional
import pandas as pd
from datetime import datetime

# Add src to path - navigate from scripts/03_sku_approval/ to project root
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))
sys.path.insert(0, str(project_root / "src"))

try:
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:
    print("❌ openpyxl not installed. Please run: pip install openpyxl>=3.1.0")
    sys.exit(1)

from src.data.database import DatabaseManager
from src.data.models import InvoiceItem, ApprovedSku
from sqlalchemy import text


class ExcelSkuApprovalManager:
    """Excel-based SKU approval workflow with formula-powered dependent dropdowns."""

    def __init__(self):
        """Initialize database connection and settings."""
        self.db_manager = DatabaseManager()
        self.approval_dir = Path("data/approval")
        self.approval_dir.mkdir(exist_ok=True)
        self.p62_categories = self._load_p62_categories()
        self.standardized_units = ["Litros", "Kilogramos", "Piezas"]

        print("🎯 Excel SKU Approval Manager initialized (Formula-Powered)")
        print(f"   📁 Approval directory: {self.approval_dir}")
        categories = self.p62_categories.get('categories', {})
        total_subcategories = sum(len(subcats) for subcats in categories.values())
        print(f"   📋 P62 Taxonomy loaded: {len(categories)} Categories, {total_subcategories} SubCategories")

    def _load_p62_categories(self) -> Dict[str, Any]:
        """Load P62 categories from configuration."""
        try:
            with open("config/p62_categories.json", 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"❌ Error loading P62 categories: {e}")
            return {"categories": {}}

    def sanitize_name(self, name: str) -> str:
        """Sanitize a string to be a valid Excel named range."""
        if not isinstance(name, str) or not name:
            return "INVALID_NAME"
        # Replace common delimiters with underscores.
        # This logic MUST be replicable in an Excel formula.
        name = name.replace(' ', '_').replace('-', '_').replace('/', '_')
        # Remove any other non-alphanumeric characters
        return re.sub(r'[^a-zA-Z0-9_]', '', name)

    def export_pending_skus(self, min_frequency: int = 1, min_value: float = 0.0) -> Optional[str]:
        """Export pending SKUs to a formula-powered Excel file."""
        print("\n🔄 EXPORTING PENDING SKUs")
        print("=" * 60)
        
        try:
            with self.db_manager.get_session() as session:
                query = text("""
                    SELECT 
                        ii.sku_key as ID,
                        ii.description as Description,
                        inv.issuer_name as Supplier,
                        ii.category as AI_Category,
                        ii.subcategory as AI_SubCategory,
                        ii.sub_sub_category as AI_SubSubCategory,
                        ii.unit_code as Original_Unit,
                        ii.standardized_unit as AI_Unit,
                        COALESCE(ii.conversion_factor, 1.0) as AI_Units_Per_Package,
                        'FALSE' as Is_General_Expense,
                        COUNT(*) as Frequency,
                        SUM(ii.total_amount) as Total_Value_MXN
                    FROM invoice_items ii
                    JOIN invoices inv ON ii.invoice_id = inv.id
                    LEFT JOIN approved_skus aps ON ii.sku_key = aps.sku_key
                    WHERE ii.approval_status = 'pending' AND aps.sku_key IS NULL AND ii.sku_key IS NOT NULL
                    GROUP BY ii.sku_key, ii.description, inv.issuer_name, ii.category, ii.subcategory, 
                             ii.sub_sub_category, ii.unit_code, ii.standardized_unit, ii.conversion_factor
                    HAVING COUNT(*) >= :min_frequency AND SUM(ii.total_amount) >= :min_value
                    ORDER BY SUM(ii.total_amount) DESC, COUNT(*) DESC
                """)
                
                result = session.execute(query, {"min_frequency": min_frequency, "min_value": min_value})
                rows = result.fetchall()
                
                if not rows:
                    print("✅ No pending SKUs found matching the criteria.")
                    return None
                
                # Convert result rows to DataFrame using column names
                columns = result.keys()
                df = pd.DataFrame([dict(zip(columns, row)) for row in rows])
                df['Priority_Score'] = (df['Frequency'] * df['Total_Value_MXN']).round(2)

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                excel_path = self.approval_dir / f"sku_approval_{timestamp}.xlsx"
                
                self._create_approval_workbook(df, excel_path)
                
                print(f"\n✅ Successfully exported {len(df)} pending SKUs.")
                print(f"📄 Excel file ready at: {excel_path}")
                print(f"💰 Total business impact of pending items: ${df['Total_Value_MXN'].sum():,.2f} MXN")
                
                return str(excel_path)
                
        except Exception as e:
            print(f"❌ Export failed: {e}")
            raise
    
    def _create_approval_workbook(self, df: pd.DataFrame, excel_path: Path) -> None:
        """Create the complete Excel workbook with data and formula-based dropdowns."""
        print("🔧 Creating formula-powered Excel workbook...")
        
        wb = openpyxl.Workbook()
        
        # 1. Create the data source sheets and all necessary named ranges
        self._create_taxonomy_sheet_and_named_ranges(wb)

        # 2. Set up the main approval sheet with new layout
        ws_main = wb.active
        ws_main.title = "SKU_Approval"
        self._setup_main_sheet(ws_main, df)

        # 3. Add formula-based data validations (dropdowns)
        self._add_data_validations(ws_main, len(df))

        # 4. Apply formatting
        self._format_sheet(ws_main)
        
        # Per user request, the helper sheets are visible.
        # wb["P62_Taxonomy"].sheet_state = 'hidden'
        # wb["Helper_Lists"].sheet_state = 'hidden'

        wb.save(excel_path)
        print(f"   💾 Workbook saved: {excel_path.name}")

    def _create_taxonomy_sheet_and_named_ranges(self, wb: openpyxl.Workbook) -> None:
        """Create sheets with taxonomy data and define all named ranges for dropdowns."""
        # Create the main taxonomy sheet
        ws_tax = wb.create_sheet("P62_Taxonomy")
        ws_tax.append(["Category", "SubCategory", "SubSubCategory"])

        # Populate the taxonomy sheet from p62_categories.json
        taxonomy_data = []
        categories_dict = self.p62_categories.get("categories", {})
        for cat, subcats in categories_dict.items():
            for subcat, subsubcats in subcats.items():
                if subsubcats:
                    for subsubcat in subsubcats:
                        taxonomy_data.append([cat, subcat, subsubcat])
                else:
                    taxonomy_data.append([cat, subcat, None])
        
        for row in taxonomy_data:
            ws_tax.append(row)

        # --- Create Named Ranges ---
        # 1. Main Categories
        categories = sorted(categories_dict.keys())
        # Create a helper sheet for lists
        ws_lists = wb.create_sheet("Helper_Lists")
        ws_lists.append(["Categories"])
        for cat in categories:
            ws_lists.append([cat])
        
        # Define named range for categories
        range_end = len(categories) + 1
        cat_range = DefinedName('Categories', attr_text=f'Helper_Lists!$A$2:$A${range_end}')
        wb.defined_names['Categories'] = cat_range

        # 2. Sub-Categories (dependent on Category)
        for cat in categories:
            subcats = sorted(categories_dict.get(cat, {}).keys())
            if subcats:
                range_name = self.sanitize_name(cat)
                # Add list to a new column in helper sheet
                col_letter = openpyxl.utils.get_column_letter(ws_lists.max_column + 1)
                ws_lists[f'{col_letter}1'] = range_name
                for i, subcat in enumerate(subcats, 2):
                    ws_lists[f'{col_letter}{i}'] = subcat
                
                range_end = len(subcats) + 1
                subcat_range = DefinedName(range_name, attr_text=f'Helper_Lists!${col_letter}$2:${col_letter}${range_end}')
                wb.defined_names[range_name] = subcat_range

        # 3. Sub-Sub-Categories (dependent on SubCategory)
        all_subcats = {sc for cat in categories_dict.values() for sc in cat.keys()}
        for subcat in sorted(list(all_subcats)):
            # Find its parent category to get the list of sub-sub-categories
            for cat_data in categories_dict.values():
                if subcat in cat_data:
                    subsubcats = sorted(cat_data.get(subcat, []))
                    if subsubcats:
                        range_name = self.sanitize_name(subcat)
                        # Add list to a new column in helper sheet
                        col_letter = openpyxl.utils.get_column_letter(ws_lists.max_column + 1)
                        ws_lists[f'{col_letter}1'] = range_name
                        for i, subsubcat in enumerate(subsubcats, 2):
                            ws_lists[f'{col_letter}{i}'] = subsubcat
                        
                        range_end = len(subsubcats) + 1
                        subsubcat_range = DefinedName(range_name, attr_text=f'Helper_Lists!${col_letter}$2:${col_letter}${range_end}')
                        wb.defined_names[range_name] = subsubcat_range
                    break # Found the subcat, no need to continue inner loop
        
        # 4. Other lists (Units, Expense Options)
        ws_lists['Z1'] = "StandardizedUnits"
        for i, unit in enumerate(self.standardized_units, 2):
            ws_lists[f'Z{i}'] = unit
        unit_range = DefinedName('StandardizedUnits', attr_text=f'Helper_Lists!$Z$2:$Z${len(self.standardized_units)+1}')
        wb.defined_names['StandardizedUnits'] = unit_range

        ws_lists['AA1'] = "ExpenseOptions"
        ws_lists['AA2'] = "TRUE"
        ws_lists['AA3'] = "FALSE"
        expense_range = DefinedName('ExpenseOptions', attr_text='Helper_Lists!$AA$2:$AA$3')
        wb.defined_names['ExpenseOptions'] = expense_range

    def _setup_main_sheet(self, ws, df: pd.DataFrame) -> None:
        """Set up the main approval sheet with the new AI vs Human layout."""
        # Title
        ws['A1'] = "SKU APPROVAL SYSTEM (Formula-Powered)"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        ws.merge_cells('A1:P1')
        
        # Headers
        headers = [
            # Reference
            "Priority Score", "ID", "Description", "Supplier", "Original Unit",
            # AI Proposed
            "AI Category", "AI SubCategory", "AI SubSubCategory", "AI Unit", "AI Units/Package",
            # Human Verified
            "Verified Category ▼", "Verified SubCategory ▼", "Verified SubSubCategory ▼",
            "Verified Unit ▼", "Verified Units/Package", "General Expense ▼"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Data columns to select and order from DataFrame
        data_columns = [
            'Priority_Score', 'ID', 'Description', 'Supplier', 'Original_Unit',
            'AI_Category', 'AI_SubCategory', 'AI_SubSubCategory', 'AI_Unit', 'AI_Units_Per_Package'
        ]
        
        # Pre-fill verified columns with AI data
        df['Verified_Category'] = df['AI_Category']
        df['Verified_SubCategory'] = df['AI_SubCategory']
        df['Verified_SubSubCategory'] = df['AI_SubSubCategory']
        df['Verified_Unit'] = df['AI_Unit']
        df['Verified_Units_Per_Package'] = df['AI_Units_Per_Package']
        df['General_Expense'] = df['Is_General_Expense']

        all_columns = data_columns + [
            'Verified_Category', 'Verified_SubCategory', 'Verified_SubSubCategory',
            'Verified_Unit', 'Verified_Units_Per_Package', 'General_Expense'
        ]

        for row_idx, row_data in enumerate(df.to_dict(orient='records'), 3):
            for col_idx, column_key in enumerate(all_columns, 1):
                value = row_data.get(column_key, '')
                ws.cell(row=row_idx, column=col_idx, value=value if pd.notna(value) else "")

    def _add_data_validations(self, ws, data_rows: int) -> None:
        """Add formula-based data validations for the dropdowns."""
        if data_rows == 0:
            return
        last_row = data_rows + 2

        # Sanitize formula for INDIRECT
        # Replicates the logic from sanitize_name() inside an Excel formula
        sanitize_formula = lambda cell: f"SUBSTITUTE(SUBSTITUTE(SUBSTITUTE({cell},\" \",\"_\"),\"-\",\"_\"),\"/\",\"_\")"

        # Verified Category (Column K)
        dv_cat = DataValidation(type="list", formula1="=Categories")
        ws.add_data_validation(dv_cat)
        dv_cat.add(f'K3:K{last_row}')

        # Verified SubCategory (Column L) - dependent on Column K
        dv_subcat = DataValidation(type="list", formula1=f"=INDIRECT({sanitize_formula('K3')})")
        ws.add_data_validation(dv_subcat)
        dv_subcat.add(f'L3:L{last_row}')
        
        # Verified SubSubCategory (Column M) - dependent on Column L
        dv_subsubcat = DataValidation(type="list", formula1=f"=INDIRECT({sanitize_formula('L3')})")
        ws.add_data_validation(dv_subsubcat)
        dv_subsubcat.add(f'M3:M{last_row}')

        # Verified Unit (Column N)
        dv_unit = DataValidation(type="list", formula1="=StandardizedUnits")
        ws.add_data_validation(dv_unit)
        dv_unit.add(f'N3:N{last_row}')

        # General Expense (Column P)
        dv_expense = DataValidation(type="list", formula1="=ExpenseOptions")
        ws.add_data_validation(dv_expense)
        dv_expense.add(f'P3:P{last_row}')

    def _format_sheet(self, ws) -> None:
        """Format the main approval sheet."""
        # Column widths
        column_widths = {
            'A': 12, 'B': 25, 'C': 40, 'D': 25, 'E': 15, # Ref
            'F': 20, 'G': 25, 'H': 25, 'I': 15, 'J': 15, # AI
            'K': 22, 'L': 25, 'M': 25, 'N': 18, 'O': 18, 'P': 15 # Verified
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze panes for scrolling
        ws.freeze_panes = 'F3'
    
    def import_approved_skus(self, excel_path: str) -> Dict[str, Any]:
        """Import approved SKU classifications from the Excel file."""
        print("\n📥 IMPORTING APPROVED SKUs FROM EXCEL")
        print("=" * 50)
        
        excel_file_path = Path(excel_path)
        if not excel_file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        try:
            df = pd.read_excel(excel_path, sheet_name='SKU_Approval', skiprows=1, keep_default_na=False)

            # Define the columns to read based on the new layout
            df.columns = [
                "Priority_Score", "ID", "Description", "Supplier", "Original_Unit",
                "AI_Category", "AI_SubCategory", "AI_SubSubCategory", "AI_Unit", "AI_Units_Per_Package",
                "Category", "SubCategory", "SubSubCategory",
                "Unit", "Units_Per_Package", "General_Expense"
            ]
            
            df = df.dropna(subset=['ID'])
            df = df[df['ID'] != ''] # Ensure no blank IDs

            stats = {'total_rows': len(df), 'successful_imports': 0, 'skipped': 0, 'errors': []}
            
            for idx, row in df.iterrows():
                try:
                    sku_key = str(row['ID']).strip()
                    with self.db_manager.get_session() as session:
                        existing = session.query(ApprovedSku).filter_by(sku_key=sku_key).first()
                        if existing:
                            stats['skipped'] += 1
                            continue

                        approved_sku = ApprovedSku(
                            sku_key=sku_key,
                            normalized_description=str(row['Description']).strip(),
                            category=str(row['Category']).strip(),
                            subcategory=str(row['SubCategory']).strip(),
                            sub_sub_category=str(row['SubSubCategory']).strip(),
                            standardized_unit=str(row['Unit']).strip(),
                            units_per_package=float(row['Units_Per_Package']),
                            approved_by='excel_import',
                            approval_date=datetime.utcnow(),
                            confidence_score=1.0  # Human approved
                        )
                        session.add(approved_sku)
                        
                        session.query(InvoiceItem).filter(InvoiceItem.sku_key == sku_key).update({
                            'approval_status': 'approved',
                            'category': approved_sku.category,
                            'subcategory': approved_sku.subcategory,
                            'sub_sub_category': approved_sku.sub_sub_category,
                            'standardized_unit': approved_sku.standardized_unit,
                            'conversion_factor': approved_sku.units_per_package,
                            'units_per_package': approved_sku.units_per_package
                        }, synchronize_session=False)
                        
                        # Also update purchase_details table for Google Sheets export
                        session.execute(
                            text("""
                                UPDATE purchase_details 
                                SET approval_status = :approval_status,
                                    category = :category,
                                    subcategory = :subcategory,
                                    sub_sub_category = :sub_sub_category,
                                    standardized_unit = :standardized_unit,
                                    conversion_factor = :conversion_factor,
                                    units_per_package = :units_per_package,
                                    updated_at = :updated_at
                                WHERE sku_key = :sku_key
                            """),
                            {
                                'approval_status': 'approved',
                                'category': approved_sku.category,
                                'subcategory': approved_sku.subcategory,
                                'sub_sub_category': approved_sku.sub_sub_category,
                                'standardized_unit': approved_sku.standardized_unit,
                                'conversion_factor': approved_sku.units_per_package,
                                'units_per_package': approved_sku.units_per_package,
                                'updated_at': datetime.utcnow(),
                                'sku_key': sku_key
                            }
                        )
                        
                        session.commit()
                        stats['successful_imports'] += 1
                except Exception as e:
                    error_msg = f"Row {idx + 3}: {e}"
                    stats['errors'].append(error_msg)
                    print(f"   ⚠️  {error_msg}")
            
            print("\n✅ Import completed:")
            print(f"   📊 Total rows processed: {stats['total_rows']}")
            print(f"   ✅ Successfully imported: {stats['successful_imports']}")
            print(f"   ⏩ Skipped (already exist): {stats['skipped']}")
            print(f"   ❌ Errors: {len(stats['errors'])}")
            
            return stats
            
        except Exception as e:
            print(f"❌ A critical error occurred during import: {e}")
            raise


def main():
    """CLI for the formula-powered Excel SKU approval workflow."""
    
    print("🎯 EXCEL SKU APPROVAL SYSTEM v5 - FORMULA-POWERED")
    print("=" * 60)
    
    manager = ExcelSkuApprovalManager()
    
    if len(sys.argv) == 1 or (len(sys.argv) > 1 and sys.argv[1] == 'export'):
        # Handle 'export' command or default action
        min_freq = int(sys.argv[2]) if len(sys.argv) > 2 and sys.argv[1] == 'export' else 1
        min_value = float(sys.argv[3]) if len(sys.argv) > 3 and sys.argv[1] == 'export' else 0.0
        
        excel_path = manager.export_pending_skus(min_freq, min_value)
        if excel_path:
            print(f"\n🎯 NEXT STEPS:")
            print(f"1. Open the Excel file: {excel_path}")
            print(f"2. Review the 'AI Proposed' columns.")
            print(f"3. Correct any values in the 'Verified' columns using the dropdowns.")
            print(f"4. Save the file and run the import command:")
            print(f'   python scripts/03_sku_approval/excel_approval.py import "{excel_path}"')
    
    elif sys.argv[1] == 'import':
        if len(sys.argv) < 3:
            print("❌ Error: Please provide the path to the Excel file for import.")
            print("Usage: python scripts/03_sku_approval/excel_approval.py import <path_to_excel_file>")
            return
        
        excel_path = sys.argv[2]
        if not Path(excel_path).exists():
            print(f"❌ Error: File not found at '{excel_path}'")
            return
        
        manager.import_approved_skus(excel_path)
    
    else:
        print("Usage:")
        print("  - To export pending SKUs (default action):")
        print("    python scripts/03_sku_approval/excel_approval.py")
        print("    python scripts/03_sku_approval/excel_approval.py export [min_frequency] [min_total_value]")
        print("\n  - To import reviewed SKUs from an Excel file:")
        print("    python scripts/03_sku_approval/excel_approval.py import <path_to_excel_file>")

if __name__ == "__main__":
    main() 